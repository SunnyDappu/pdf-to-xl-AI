"""
Excel Generator Service
Converts extracted data into formatted Excel workbooks.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import Dict, List
import os

class ExcelGenerator:
    def __init__(self):
        self.fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.font_header = Font(bold=True, color="FFFFFF", size=12)
        self.alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    def generate_excel(self, merged_data: Dict, customer_request: str, output_path: str) -> bool:
        """
        Generate Excel file from merged chunk data.
        
        Args:
            merged_data: Result from claude_processor.merge_results()
            customer_request: Original customer request string
            output_path: Path to save Excel file
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            wb = Workbook()
            
            # Create metadata worksheet
            self._create_metadata_sheet(wb, merged_data, customer_request)
            
            # Create data worksheet
            self._create_data_sheet(wb, merged_data)
            
            # Save workbook
            wb.save(output_path)
            return True
        
        except Exception as e:
            print(f"Error generating Excel: {e}")
            return False
    
    def _create_metadata_sheet(self, wb: Workbook, merged_data: Dict, customer_request: str):
        """Create metadata worksheet with processing information"""
        ws = wb.active
        ws.title = "Processing Info"
        
        # Title
        ws['A1'] = "PDF-to-Excel Processing Report"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Metadata
        row = 3
        metadata_items = [
            ("Customer Request:", customer_request),
            ("Processing Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Chunks Processed:", merged_data['quality_metrics']['total_chunks']),
            ("Successful Chunks:", merged_data['quality_metrics']['successful_chunks']),
            ("Failed Chunks:", merged_data['quality_metrics']['failed_chunks']),
            ("Total Records Extracted:", merged_data['quality_metrics']['total_records']),
            ("Fields Extracted:", merged_data['quality_metrics']['field_count']),
            ("Field Names:", ", ".join(merged_data['field_schema'])),
        ]
        
        for label, value in metadata_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)
            row += 1
        
        # Error log if any
        if merged_data['chunk_errors']:
            row += 1
            ws[f'A{row}'] = "Chunk Processing Errors:"
            ws[f'A{row}'].font = Font(bold=True, color="FF0000")
            row += 1
            
            for error in merged_data['chunk_errors']:
                ws[f'A{row}'] = f"Chunk {error['chunk_id']}: {error['error']}"
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_data_sheet(self, wb: Workbook, merged_data: Dict):
        """Create main data worksheet with extracted information"""
        ws = wb.create_sheet("Extracted Data")
        
        field_schema = merged_data['field_schema']
        data = merged_data['merged_data']
        
        # Write headers
        for col_idx, field_name in enumerate(field_schema, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.alignment_center
        
        # Write data rows
        for row_idx, record in enumerate(data, 2):
            for col_idx, field_name in enumerate(field_schema, 1):
                value = record.get(field_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.alignment_left
                
                # Color alternate rows for readability
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Adjust column widths
        for col_idx, field_name in enumerate(field_schema, 1):
            max_length = len(str(field_name))
            for row_idx, record in enumerate(data, 2):
                value = record.get(field_name, '')
                max_length = max(max_length, len(str(value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def save_excel(self, output_path: str) -> str:
        """Save Excel file and return the path"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        return output_path
